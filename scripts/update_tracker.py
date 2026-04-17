"""
Update the job application tracker Excel spreadsheet.

Usage:
    python update_tracker.py <tracker.xlsx> --add <application.json>
    python update_tracker.py <tracker.xlsx> --update <row_number> <field> <value>

application.json schema:
{
    "company": "Spotify",
    "role": "PM - Artist Promotion, Payments & Access",
    "type": "PM (Monetization)",
    "date_applied": "Not yet",
    "status": "Prepping",
    "resume_path": "spotify-artist-promotion-pm/Anisha_Subberwal_Resume_Spotify_PM.pdf",
    "cover_letter_path": "spotify-artist-promotion-pm/CoverLetter_Spotify.pdf",
    "outreach": "Yes",
    "referral": "",
    "comp": "$96K-$137K + equity",
    "notes": "Music Mission / Spotify for Artists team.",
    "url": "https://linkedin.com/jobs/view/..."
}
"""

import json
import sys
import os
import argparse
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# Color palette
HEADER_BG = "1B2A4A"       # Dark navy
HEADER_FG = "FFFFFF"        # White text
ROW_EVEN = "F7F8FC"         # Light blue-gray
ROW_ODD = "FFFFFF"          # White
ACCENT = "4A7BF7"           # Blue accent
STATUS_COLORS = {
    "Prepping":       ("FFF3E0", "E65100"),   # Orange bg/text
    "Applied":        ("E3F2FD", "1565C0"),   # Blue bg/text
    "Outreach Sent":  ("F3E5F5", "7B1FA2"),   # Purple bg/text
    "Interviewing":   ("E8F5E9", "2E7D32"),   # Green bg/text
    "Offer":          ("C8E6C9", "1B5E20"),   # Dark green
    "Rejected":       ("FFEBEE", "C62828"),    # Red
    "Ghosted":        ("ECEFF1", "546E7A"),    # Gray
}

COLUMNS = [
    ("#", 5),
    ("Company", 20),
    ("Role", 38),
    ("Type", 18),
    ("Date Applied", 14),
    ("Status", 15),
    ("Resume", 12),
    ("Cover Letter", 14),
    ("Outreach", 11),
    ("Referral", 18),
    ("Comp", 22),
    ("Notes", 35),
    ("URL", 30),
]

thin_border = Border(
    left=Side(style='thin', color='D0D5DD'),
    right=Side(style='thin', color='D0D5DD'),
    top=Side(style='thin', color='D0D5DD'),
    bottom=Side(style='thin', color='D0D5DD'),
)


def create_workbook(path):
    """Create a new tracker workbook with formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Header row
    header_font = Font(name='Calibri', size=11, bold=True, color=HEADER_FG)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Set row height for header
    ws.row_dimensions[1].height = 30

    # Add conditional formatting for status column (column 6)
    for status, (bg, fg) in STATUS_COLORS.items():
        ws.conditional_formatting.add(
            f'F2:F100',
            CellIsRule(
                operator='equal',
                formula=[f'"{status}"'],
                fill=PatternFill(start_color=bg, end_color=bg, fill_type='solid'),
                font=Font(color=fg, bold=True, size=10)
            )
        )

    # Add a summary sheet
    summary = wb.create_sheet("Summary")
    summary['A1'] = "Job Application Summary"
    summary['A1'].font = Font(name='Calibri', size=14, bold=True, color=HEADER_BG)

    summary['A3'] = "Total Applications"
    summary['B3'] = '=COUNTA(Applications!A2:A100)'
    summary['A4'] = "Applied"
    summary['B4'] = '=COUNTIF(Applications!F2:F100,"Applied")'
    summary['A5'] = "Interviewing"
    summary['B5'] = '=COUNTIF(Applications!F2:F100,"Interviewing")'
    summary['A6'] = "Offers"
    summary['B6'] = '=COUNTIF(Applications!F2:F100,"Offer")'
    summary['A7'] = "Prepping"
    summary['B7'] = '=COUNTIF(Applications!F2:F100,"Prepping")'
    summary['A8'] = "Outreach Sent"
    summary['B8'] = '=COUNTIF(Applications!F2:F100,"Outreach Sent")'
    summary['A9'] = "Rejected"
    summary['B9'] = '=COUNTIF(Applications!F2:F100,"Rejected")'
    summary['A10'] = "Ghosted"
    summary['B10'] = '=COUNTIF(Applications!F2:F100,"Ghosted")'

    for row in range(3, 11):
        summary.cell(row=row, column=1).font = Font(name='Calibri', size=11)
        summary.cell(row=row, column=2).font = Font(name='Calibri', size=11, bold=True)
        summary.cell(row=row, column=2).alignment = Alignment(horizontal='center')

    summary['A3'].font = Font(name='Calibri', size=11, bold=True)
    summary.column_dimensions['A'].width = 22
    summary.column_dimensions['B'].width = 12

    # Status breakdown header styling
    for row in [4, 5, 6, 7, 8, 9, 10]:
        label = summary.cell(row=row, column=1).value
        if label in STATUS_COLORS:
            bg, fg = STATUS_COLORS[label]
            summary.cell(row=row, column=2).fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            summary.cell(row=row, column=2).font = Font(name='Calibri', size=11, bold=True, color=fg)

    wb.save(path)
    return wb


def add_application(path, app_data):
    """Add a new application row to the tracker."""
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = create_workbook(path)
        wb = load_workbook(path)

    ws = wb['Applications']

    # Find next row
    next_row = ws.max_row + 1
    row_num = next_row - 1  # Application number

    # Determine row color
    row_fill = PatternFill(
        start_color=ROW_EVEN if row_num % 2 == 0 else ROW_ODD,
        end_color=ROW_EVEN if row_num % 2 == 0 else ROW_ODD,
        fill_type='solid'
    )

    cell_font = Font(name='Calibri', size=10)
    cell_align = Alignment(vertical='center', wrap_text=True)

    values = [
        row_num,
        app_data.get('company', ''),
        app_data.get('role', ''),
        app_data.get('type', ''),
        app_data.get('date_applied', 'Not yet'),
        app_data.get('status', 'Prepping'),
        app_data.get('resume_path', ''),
        app_data.get('cover_letter_path', ''),
        app_data.get('outreach', ''),
        app_data.get('referral', ''),
        app_data.get('comp', ''),
        app_data.get('notes', ''),
        app_data.get('url', ''),
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        cell.font = cell_font
        cell.fill = row_fill
        cell.alignment = cell_align
        cell.border = thin_border

    # Bold the company name
    ws.cell(row=next_row, column=2).font = Font(name='Calibri', size=10, bold=True)

    # Make URL a hyperlink if present
    url = app_data.get('url', '')
    if url:
        cell = ws.cell(row=next_row, column=13)
        cell.hyperlink = url
        cell.font = Font(name='Calibri', size=10, color=ACCENT, underline='single')

    # Set row height
    ws.row_dimensions[next_row].height = 28

    wb.save(path)
    print(f'Added row {row_num}: {app_data.get("company")} — {app_data.get("role")}')


def update_field(path, row_num, field, value):
    """Update a specific field in an existing application row."""
    wb = load_workbook(path)
    ws = wb['Applications']

    field_map = {
        'company': 2, 'role': 3, 'type': 4, 'date_applied': 5,
        'status': 6, 'resume': 7, 'cover_letter': 8, 'outreach': 9,
        'referral': 10, 'comp': 11, 'notes': 12, 'url': 13,
    }

    if field not in field_map:
        print(f'Unknown field: {field}. Options: {", ".join(field_map.keys())}')
        sys.exit(1)

    excel_row = row_num + 1  # +1 for header
    col = field_map[field]
    ws.cell(row=excel_row, column=col, value=value)

    wb.save(path)
    print(f'Updated row {row_num} — {field}: {value}')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('tracker', help='Path to tracker.xlsx')
    sub = parser.add_subparsers(dest='command')

    add_p = sub.add_parser('add')
    add_p.add_argument('data', help='Path to application JSON')

    upd_p = sub.add_parser('update')
    upd_p.add_argument('row', type=int, help='Row number to update')
    upd_p.add_argument('field', help='Field name to update')
    upd_p.add_argument('value', help='New value')

    init_p = sub.add_parser('init')

    args = parser.parse_args()

    if args.command == 'add':
        with open(args.data, 'r', encoding='utf-8') as f:
            data = json.load(f)
        add_application(args.tracker, data)
    elif args.command == 'update':
        update_field(args.tracker, args.row, args.field, args.value)
    elif args.command == 'init':
        create_workbook(args.tracker)
        print(f'Created tracker: {args.tracker}')
    else:
        parser.print_help()
